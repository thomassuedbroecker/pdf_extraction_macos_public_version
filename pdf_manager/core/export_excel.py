"""Excel export for visible PDF table records."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pdf_manager.models.pdf_record import PdfRecord

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)

COLUMN_LABELS = {
    "file_name": "File Name",
    "full_path": "Full Path",
    "parent_folder": "Parent Folder",
    "file_size_bytes": "File Size Bytes",
    "file_size_mb": "File Size MiB",
    "created_date": "Created Date",
    "modified_date": "Modified Date",
    "page_count": "Pages",
    "title": "Title",
    "author": "Author",
    "subject": "Subject",
    "producer": "Producer/Application",
    "encrypted": "Encrypted",
    "text_extractable": "Text Extractable",
    "text_preview": "Text Preview",
    "error": "Error",
}


def export_records_to_excel(
    records: Iterable[PdfRecord],
    output_path: str | Path,
    columns: list[str],
    scanned_roots: list[str],
) -> Path:
    """Export records to a formatted Excel workbook."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PDFs"

    headers = [COLUMN_LABELS.get(column, column) for column in columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for record in records:
        data = record.to_dict()
        sheet.append([data.get(column, "") for column in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 70)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    metadata = workbook.create_sheet("Export Metadata")
    metadata.append(["Export Date", datetime.now().isoformat(sep=" ", timespec="seconds")])
    metadata.append(["Scanned Root Folders", ""])
    for root in scanned_roots:
        metadata.append(["", root])
    metadata.append(["Visible Columns", ", ".join(headers)])
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 90
    for cell in metadata[1]:
        cell.font = HEADER_FONT

    workbook.save(output)
    return output


def export_extraction_results_to_excel(
    results: Iterable[dict[str, str]],
    output_path: str | Path,
) -> Path:
    """Export local LLM extraction results to a formatted Excel workbook."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LLM Results"

    headers = ["File", "Path", "Status", "Result / Error"]
    keys = ["file_name", "full_path", "status", "result"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    row_count = 0
    for row in results:
        sheet.append([row.get(key, "") for key in keys])
        row_count += 1

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 90)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    metadata = workbook.create_sheet("Export Metadata")
    metadata.append(["Export Date", datetime.now().isoformat(sep=" ", timespec="seconds")])
    metadata.append(["Result Rows", row_count])
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 90
    for cell in metadata[1]:
        cell.font = HEADER_FONT

    workbook.save(output)
    return output
