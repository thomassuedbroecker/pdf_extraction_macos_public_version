from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pdf_manager.core.export_excel import export_records_to_excel
from pdf_manager.models.pdf_record import PdfRecord


pytestmark = pytest.mark.unit
COLUMNS = ["file_name", "full_path", "file_size_mb", "page_count", "title", "encrypted"]


def test_export_excel_creates_xlsx_file(sample_pdf_records: list[PdfRecord], temporary_export_path: Path) -> None:
    export_records_to_excel(sample_pdf_records, temporary_export_path, COLUMNS, ["/tmp"])

    assert temporary_export_path.exists()
    assert load_workbook(temporary_export_path)["PDFs"].max_row == 3


def test_export_excel_contains_expected_headers(sample_pdf_records: list[PdfRecord], temporary_export_path: Path) -> None:
    export_records_to_excel(sample_pdf_records, temporary_export_path, COLUMNS, ["/tmp"])

    sheet = load_workbook(temporary_export_path)["PDFs"]

    assert [cell.value for cell in sheet[1]] == [
        "File Name",
        "Full Path",
        "File Size MiB",
        "Pages",
        "Title",
        "Encrypted",
    ]


def test_export_excel_exports_filtered_records_only(
    sample_pdf_records: list[PdfRecord], temporary_export_path: Path
) -> None:
    export_records_to_excel([sample_pdf_records[0]], temporary_export_path, COLUMNS, ["/tmp"])

    sheet = load_workbook(temporary_export_path)["PDFs"]

    assert sheet.max_row == 2
    assert sheet["A2"].value == "alpha.pdf"


def test_export_excel_adds_metadata_sheet(sample_pdf_records: list[PdfRecord], temporary_export_path: Path) -> None:
    export_records_to_excel(sample_pdf_records, temporary_export_path, COLUMNS, ["/tmp/root"])

    workbook = load_workbook(temporary_export_path)
    metadata = workbook["Export Metadata"]

    assert "Export Metadata" in workbook.sheetnames
    assert metadata["A1"].value == "Export Date"
    assert metadata["A2"].value == "Scanned Root Folders"
    assert metadata["B3"].value == "/tmp/root"
